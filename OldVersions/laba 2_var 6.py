import openpyxl
import os
import random

def generate_random_graphs(num_nodes, num_edges, filename = "input_graph.xlsx"):
    if os.path.exists(filename):
        os.remove(filename)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # nodes = [str(i) for i in range(num_nodes)]
    nodes = [chr(ord("A")+i) for i in range(num_nodes)]
    edges = {}

    while len(edges) < num_edges:
        from_node = random.choice(nodes)
        to_node = random.choice(nodes)
        if from_node != to_node: 
            weight = random.randint(1, 10)
            edges[from_node, to_node] = weight
    
    sheet.cell(row=1, column=1, value=num_nodes) 
    sheet.cell(row=1, column=2, value=num_edges) 

    for row, edge in enumerate(edges.items()): 
        (from_node, to_node), weight = edge
        sheet.cell(row=row+2, column=1, value=from_node) 
        sheet.cell(row=row+2, column=2, value=to_node) 
        sheet.cell(row=row+2, column=3, value=weight) 

    for row, node in enumerate(nodes):
        sheet.cell(row=row + 2, column=5, value=node)

    workbook.save(filename)

def read_graph_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    edges = {}
    num_nodes = sheet.cell(row=1, column=1).value
    num_edges = sheet.cell(row=1, column=2).value

    for row in range(2, num_edges + 1):
        from_node = sheet.cell(row=row, column=1).value
        to_node = sheet.cell(row=row, column=2).value
        weight = sheet.cell(row=row, column=3).value
        
        edges[(from_node, to_node)] = weight
    
    nodes = {}
    for row in range(num_nodes): 
        node = sheet.cell(row=row+2, column=5).value 
        nodes[node] = row

    workbook.close()
    graph = [[float('inf')] * num_nodes for _ in range(num_nodes)]

    for i in range(num_nodes):
        graph[i][i] = 0
    for (from_node, to_node), weight in edges.items():
        graph[nodes[from_node]][nodes[to_node]] = weight

    return graph, list(nodes.keys())

def floyd_warshall_with_path(graph):
    num_vertices = len(graph)
    dist = [row[:] for row in graph]
    next_node = [[None if i != j and graph[i][j] == float('inf') else j for j in range(num_vertices)] for i in range(num_vertices)]

    for k in range(num_vertices):
        for i in range(num_vertices):
            for j in range(num_vertices):
                if dist[i][j] > dist[i][k] + dist[k][j]:
                    dist[i][j] = dist[i][k] + dist[k][j]
                    next_node[i][j] = next_node[i][k]

    return dist, next_node

def reconstruct_path(next_node, start, end):
    if next_node[start][end] is None:
        return [] 
    path = []
    while start != end:
        path.append(start)
        start = next_node[start][end]
    path.append(end)
    return path

def write_result_to_excel(dist, nodes, next_node, filename='graph.xlsx'):
    wb = openpyxl.Workbook()
    
    ws_dist = wb.active
    ws_dist.title = "Distances"

    ws_dist.cell(row=1, column=1, value='из\\до')
    for col, node in enumerate(nodes, start=2):
        ws_dist.cell(row=1, column=col, value=node)

    for i in range(len(dist)):
        ws_dist.cell(row=i + 2, column=1, value=nodes[i])
        for j in range(len(dist[i])):
            ws_dist.cell(row=i + 2, column=j + 2, value=dist[i][j])  

    ws_paths = wb.create_sheet(title="Paths")

    ws_paths.cell(row=1, column=1, value='Из')
    ws_paths.cell(row=1, column=2, value='До')
    ws_paths.cell(row=1, column=3, value='Путь')

    row = 2
    for i in range(len(nodes)):
        for j in range(len(nodes)):
            if i != j:
                path = reconstruct_path(next_node, i, j)
                ws_paths.cell(row=row, column=1, value=nodes[i]) 
                ws_paths.cell(row=row, column=2, value=nodes[j]) 
                ws_paths.cell(row=row, column=3, value=' -> '.join(nodes[k] for k in path)) 
                row += 1

    wb.save(filename)

def print_graph(dist, nodes):
    n = len(dist)
    for i in range(n):
        if i == 0:
            print("из\\до", end="\t") 
            for node in nodes:
                print(node, end="\t") 
            print()
        for j in range(n):
            if j == 0:
                print(nodes[i], end="\t") 
            print(dist[i][j], end="\t") 
        print()
    print()

def main():
    print("Выберите режим работы \n1. Сгенерировать рандомный граф с указанным количеством графов и путей \n2. Провести вычисления созданного графа !!!Внимание ответ будет записан в excel файл \n ")
    choose = int(input())

    if choose == 1:
        print("Введите количество узлов")
        num_nodes = max(int(input()), 1)
        print("Введите количество ветвей")
        num_edges = max(min(int(input()), num_nodes**2-num_nodes), 0)
        generate_random_graphs(num_nodes, num_edges)
    else:
        input_file = "input_graph.xlsx"
        output_file = "output_distances_and_paths.xlsx"
        graph, nodes = read_graph_from_excel(input_file)
        print("изначальный граф:")
        print_graph(graph, nodes)
        dist, next_node = floyd_warshall_with_path(graph)
        print("Граф минимальных расстояний:")
        print_graph(dist, nodes)
        for i in range(len(nodes)):
            for j in range(len(nodes)):
                if i != j:
                    path = reconstruct_path(next_node, i, j)
                    print(f"Путь из {nodes[i]} в {nodes[j]}:", path)
        write_result_to_excel(dist, nodes, next_node, output_file)
        print("Результаты сохранены в", output_file)

# Вызов функции main
if __name__ == "__main__":
    main()



